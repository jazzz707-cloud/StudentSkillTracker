from flask import Flask, render_template, request, redirect
import openpyxl
import os

app = Flask(__name__)

# Excel setup
FILE = "students.xlsx"

if not os.path.exists(FILE):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Students"
    sheet.append(["Roll No", "Name", "Department", "Skill ID", "Skill", "Achievement"])
    wb.save(FILE)

# Home page (form)
@app.route("/")
def home():
    return render_template("index.html")

# ✅ FIX: Added the route decorator here
@app.route("/save", methods=["POST"])
def save():
    roll = request.form["roll"]
    name = request.form["name"]
    dept = request.form["dept"]
    skill = request.form["skill"]
    achievement = request.form["achievement"]

    # Generate Skill ID
    skill_id = f"SKL{str(abs(hash(skill + roll)))[0:5]}"

    wb = openpyxl.load_workbook(FILE)
    sheet = wb.active
    sheet.append([roll, name, dept, skill_id, skill, achievement])
    wb.save(FILE)

    return redirect("/display")

# Display all records
@app.route("/display")
def display():
    wb = openpyxl.load_workbook(FILE)
    sheet = wb.active
    data = list(sheet.iter_rows(values_only=True))[1:]  # skip header
    return render_template("display.html", students=data)

if __name__ == "__main__":
    app.run(debug=True)
